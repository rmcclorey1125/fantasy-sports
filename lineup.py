import pandas as pd
from pulp import *
import openpyxl
import re

players = pd.read_csv(r"/Users/ronan/Downloads/FanDuel-NFL-2020-11-15-51566-players-list.csv", usecols= ['Id', 'Position', 'FPPG', 'Salary'])

wb = openpyxl.Workbook()
ws = wb.active

availables = players.groupby(["Position", "Id", "FPPG", "Salary"]).agg('count')
availables = availables.reset_index()

salaries = {}
points = {}

for pos in availables.Position.unique():
    available_pos = availables[availables.Position == pos]
    salary = list(available_pos[['Id', 'Salary']].set_index("Id").to_dict().values())[0]
    point = list(available_pos[['Id', 'FPPG']].set_index("Id").to_dict().values())[0]

    salaries[pos] = salary
    points[pos] = point

pos_num_available = {
    "QB": 1,
    "RB": 3,
    "WR": 3,
    "TE": 1,
    "DEF": 1
}

salary_cap = 60000

for lineup in range(1,51):
    _vars = {k: LpVariable.dict(k, v, cat='Binary') for k, v in points.items()}

    prob = LpProblem("Fantasy", LpMaximize)
    rewards = []
    costs = []
    position_constraints = []

    for k, v in _vars.items():
        costs += lpSum([salaries[k][i] * _vars[k][i] for i in v])
        rewards += lpSum([points[k][i] * _vars[k][i] for i in v])
        prob += lpSum([_vars[k][i] for i in v]) == pos_num_available[k]

    prob += lpSum(rewards)
    prob += lpSum(costs <= salary_cap)
    if not lineup == 1:
        prob += (lpSum(rewards) <= total_score-0.001)
    prob.solve()

    score = str(prob.objective)
    constraints = [str(const) for const in prob.constraints.values()]
    colnum = 1

    for v in prob.variables():
        score = score.replace(v.name, str(v.varValue))
        if v.varValue !=0:
            ws.cell(row=lineup, column=colnum).value = v.name
            colnum +=1
        total_score = eval(score)
        ws.cell(row=lineup, column=colnum).value = total_score
        print(lineup, total_score)

